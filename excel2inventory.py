#!/usr/bin/env python

import openpyxl
from openpyxl import load_workbook

excel_file = "inventory_updated.xlsx"
inventory_file = "ansible_inventory.ini"
wb = load_workbook(excel_file)
ws = wb.active

linux_hosts = []
windows_hosts = []

for row in ws.iter_rows(min_row=2, values_only=True):
    Server_Name, Cancellation_Status, fqdn, IP_Address, Change_Ticket, Change_Status, Patching_Group, Prepatch_Ping_Status, Prepatch_IVP, Prepatch_Reboot, In_Tanium, Postpatch_IVP, Postpatch_Reboot, Patching_Completion_Status, Patcher, Admin_Comment, Patcher_Comment, RCA_Comment, Incident, OS, Location, Country, Is_Virtual, Server_Status, ip_address, host_type, region, os_version = row
    host_entry = f"{ip_address} region={region} os_version={os_version}"
    
    if host_type.lower() == "linux":
        linux_hosts.append(host_entry)
    elif host_type.lower() == "windows":
        windows_hosts.append(host_entry)

with open(inventory_file, "w") as inventory_file:
    inventory_file.write("[Linux]\n")
    for host in linux_hosts:
        inventory_file.write(host + "\n")
    inventory_file.write("\n[Windows]\n")
    for host in windows_hosts:
        inventory_file.write(host + "\n")
